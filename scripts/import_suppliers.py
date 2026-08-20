"""
import_suppliers.py
--------------------
Reads the ERP suppliers xlsx and imports into the Supplier model.
Imports: code (Κωδικός), tin (ΑΦΜ), name (Όνομα), payment_terms (Όνομα - Τρόπος Πληρωμής).
All other fields (contacts, address, notes) remain blank for manual entry later.

use file from Pylon as exported:
    Ενέργειες / Εξαγόμενα Grid / Εξαγωγή σε Excel

Usage:
    python scripts/import_suppliers.py suppliers_list-2026-08-20.xlsx [--dry-run]

Run against Railway:
    $env:DATABASE_URL="postgresql://..."
    python scripts/import_suppliers.py suppliers_list-2026-08-20.xlsx
    $env:DATABASE_URL=""
"""

import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')

import django
django.setup()

import openpyxl
from inventory.models import Supplier

DRY_RUN = '--dry-run' in sys.argv


def clean(val):
    if val is None:
        return ''
    return str(val).strip()


def main():
    if len(sys.argv) < 2 or sys.argv[1].startswith('--'):
        print("Usage: python scripts/import_suppliers.py <file.xlsx> [--dry-run]")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        sys.exit(1)

    print(f"{'[DRY RUN] ' if DRY_RUN else ''}Reading: {filepath}\n")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Read header row
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Columns: {headers}\n")

    created = updated = skipped = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = dict(zip(headers, row))

        code          = clean(row_dict.get('Κωδικός'))
        name          = clean(row_dict.get('Όνομα'))
        tin           = clean(row_dict.get('ΑΦΜ'))
        payment_terms = clean(row_dict.get('Όνομα - Τρόπος Πληρωμής'))

        if not name:
            print(f"  Row {i}: missing name — skipped")
            skipped += 1
            continue

        if DRY_RUN:
            print(f"  DRY RUN | {code:<25} | {tin:<15} | {payment_terms:<30} | {name}")
            continue

        defaults = {'name': name, 'tin': tin, 'payment_terms': payment_terms}
        if code:
            obj, was_created = Supplier.objects.update_or_create(
                code=code, defaults=defaults
            )
        else:
            obj, was_created = Supplier.objects.update_or_create(
                name=name, defaults={**defaults, 'code': None}
            )

        if was_created:
            created += 1
            print(f"  CREATED: {code or '(no code)'} | {name}")
        else:
            updated += 1
            print(f"  UPDATED: {code or '(no code)'} | {name}")

    wb.close()

    if DRY_RUN:
        print(f"\nDry run complete — no changes made.")
    else:
        print(f"\nDone — {created} created, {updated} updated, {skipped} skipped")


if __name__ == '__main__':
    main()